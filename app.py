from flask import Flask, request, jsonify
import os
from openpyxl.drawing.image import Image
import openpyxl
from openpyxl import Workbook, load_workbook
from flask_cors import CORS
# from filelock import FileLock
from werkzeug.utils import secure_filename

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})

UPLOAD_FOLDER = "temp_images"
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/', methods=['GET'])
def hello():
    return 'hello'

@app.route('/api/sleep_data', methods=['POST'])
def post_data_excel():
    
    data = request.get_json()
    EMPLOYEE_ID_NO = data.get('EMPLOYEE_ID_NO')
    FIRST_NAME = data.get('FIRST_NAME')
    LAST_NAME = data.get('LAST_NAME')
    DEPARTMENT = data.get('DEPARTMENT')
    CONTACT = data.get('CONTACT')
    ADDRESS = data.get('ADDRESS')
    EMERGENCY_CONTACT = data.get('EMERGENCY_CONTACT')
    RELATIONSHIP = data.get('RELATIONSHIP')
    RELATIONSHIP_CONTACT = data.get('RELATIONSHIP_CONTACT')
    DATE = data.get('DATE')
    REASONS = data.get('REASONS')
    STAFF_NAME = data.get('STAFF_NAME')
    SLEEP_DATE = data.get('SLEEP_DATE')
    SUPERVISOR_NAME = data.get('SUPERVISOR_NAME')
    DESIGNATION = data.get('DESIGNATION')
    APPROVED_DATE = data.get('APPROVED_DATE')
    # ID_UPLOAD = data.get('ID_UPLOAD')
    # ID_UPLOAD = request.files['ID_UPLOAD']
    # ID_UPLOAD = Image(ID_UPLOAD)
    
   
    
    file_path = "C:\\Users\modinat.a\\Documents\\PERMISSION TO SLEEP REG DATA.xlsx"

    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    
    next_row = worksheet.max_row + 1
    worksheet.cell(row=next_row, column=1, value=EMPLOYEE_ID_NO)
    worksheet.cell(row=next_row, column=2, value=FIRST_NAME)
    worksheet.cell(row=next_row, column=3, value=LAST_NAME)
    worksheet.cell(row=next_row, column=4, value=DEPARTMENT)
    worksheet.cell(row=next_row, column=5, value=CONTACT)
    worksheet.cell(row=next_row, column=6, value=ADDRESS)
    worksheet.cell(row=next_row, column=7, value=EMERGENCY_CONTACT)
    worksheet.cell(row=next_row, column=8, value=RELATIONSHIP)
    worksheet.cell(row=next_row, column=9, value=RELATIONSHIP_CONTACT)
    worksheet.cell(row=next_row, column=10, value=DATE)
    worksheet.cell(row=next_row, column=11, value=REASONS)
    worksheet.cell(row=next_row, column=12, value=STAFF_NAME)
    worksheet.cell(row=next_row, column=13, value=SLEEP_DATE)
    worksheet.cell(row=next_row, column=14, value=SUPERVISOR_NAME)
    worksheet.cell(row=next_row, column=15, value=DESIGNATION)
    worksheet.cell(row=next_row, column=16, value=APPROVED_DATE)
    # worksheet.cell(row=next_row, column=17, value=ID_UPLOAD)
    
    workbook.save(file_path)

    return jsonify({'msg': 'Form data submitted successfully!'}) 
     





if __name__ == '__main__':
    app.run(debug=True, port=5000)


